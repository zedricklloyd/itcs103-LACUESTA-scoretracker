import tkinter as tk
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook

workbook = Workbook()
sheet = workbook.active  

sheet["A1"] = "Surname"
sheet["B1"] = "Subject"
sheet["C1"] = "Grade"
sheet["D1"] = "Result"

try:
    workbook = load_workbook("student_scores.xlsx")
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Surname"
    sheet["B1"] = "Subject"
    sheet["C1"] = "Grade"
    sheet["D1"] = "Result"
    workbook.save("student_scores.xlsx")

def save_record():
    surname = surname_entry.get()
    subject = subject_entry.get()
    grade = int(score_entry.get())
    if grade >= 1 and grade <= 74:
        result = "Fail"
        
    elif grade >= 75 and grade <= 100:
        result = "Pass"
       
    else:
        messagebox.showerror("Invalid input", "Grade must be a number.")
        return
    
    next_row = sheet.max_row + 1
    sheet[f"A{next_row}"] = surname
    sheet[f"B{next_row}"] = subject
    sheet[f"C{next_row}"] = grade
    sheet[f"D{next_row}"] = result
    workbook.save("student_scores.xlsx")

    messagebox.showinfo("Success", "Record saved!")
    surname_entry.delete(0,tk.END)
    subject_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)


window = tk.Tk()
window.title("Score Tracker")
window.geometry("300x200")

frame = tk.Frame(window)
frame.pack(expand=True)

tk.Label(frame, text="Surname:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
surname_entry = tk.Entry(frame)
surname_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")

tk.Label(frame, text="Subject:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
subject_entry = tk.Entry(frame)
subject_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

tk.Label(frame, text="Grade:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
score_entry = tk.Entry(frame)
score_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")


tk.Button(frame, text="Save Record", command=save_record).grid(row=3, column=0, columnspan=2, pady=20)


window.mainloop()
