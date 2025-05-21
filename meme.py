import tkinter as tk
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook

workbook = Workbook()
sheet = workbook.active  

sheet["A1"] = "Name"
sheet["B1"] = "Course"
sheet["C1"] = "Grade"

try:
    workbook = load_workbook("grades.xlsx")
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Course"
    sheet["C1"] = "Grade"
    workbook.save("grades.xlsx")

def save_record():
    name = name_entry.get()
    course = course_entry.get()
    grade = int(grade_entry.get())
    
    sheet.append([name, course, grade])
    workbook.save("grades.xlsx")

    messagebox.showinfo("Success", "Record saved!")
    name_entry.delete(0,tk.END)
    course_entry.delete(0, tk.END)
    grade_entry.delete(0, tk.END)

def show_data():
    
    workbook.save("grades.xlsx")

    data_window = tk.Toplevel(window)
    data_window.title("Student Data")
    data_window.geometry("300x300")

    row_number = 0
    for row in sheet.iter_rows(values_only=True):
        column_number = 0
        for cell_value in row:
           
            label = tk.Label(data_window, text=cell_value)
            label.grid(row=row_number, column=column_number, padx=5, pady=5)
            column_number += 1
        row_number += 1


window = tk.Tk()
window.title("Grade Report")
window.geometry("300x200")

frame = tk.Frame(window)
frame.pack(expand=True)

tk.Label(frame, text="Name:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
name_entry = tk.Entry(frame)
name_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")

tk.Label(frame, text="Course:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
course_entry = tk.Entry(frame)
course_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

tk.Label(frame, text="Grade:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
grade_entry = tk.Entry(frame)
grade_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")

tk.Button(frame, text="Save Record", command=save_record).grid(row=3, column=0, columnspan=2, pady=20)
tk.Button(frame, text="View Data",command=show_data).grid(row=3, column=0, columnspan=2, pady=20, sticky="e")

window.mainloop()